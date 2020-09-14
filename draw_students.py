import time
from random import sample

from openpyxl import load_workbook, Workbook


class Student:
    def __init__(self, time, email, name, parent_name, parent_number, school, grade, robotic_course,
                 student_facilities):
        self.time = time
        self.email = email
        self.name = name
        self.parent_name = parent_name
        self.parent_number = parent_number
        self.school = school
        self.grade = grade[:1]
        self.robotic_course = robotic_course
        self.student_facilities = student_facilities

    def __str__(self):
        return f"{self.name} {self.grade}. Sınıf / {self.facility_note()}"

    def is_completed_course(self):
        if self.robotic_course == "Evet":
            bool_val = True
        else:
            bool_val = False

        return bool_val

    def facility_note(self):
        facilities = self.student_facilities.split(",")
        facilities = [facility.strip() for facility in facilities]
        note = 0
        if "Bilgisayar" in facilities:
            note += 30
        if "İnternet" in facilities:
            note += 9
        if "Öğrenciye ait oda" in facilities:
            note += 7
        if "Kulaklıklı mikrofon" in facilities:
            note += 5
        if "Kulaklık" in facilities:
            note += 3
        if "Mikrofon" in facilities:
            note += 1
        return note

    def to_text(self):
        return f"{self.name}"


def create_students_list(filename):
    wb = load_workbook(filename=filename)
    ws = wb[wb.sheetnames[0]]
    students = list()

    for row in range(2, ws.max_row + 1):
        new_student = Student(time=ws[f"A{row}"].value, email=ws[f"B{row}"].value,
                              name=ws[f"C{row}"].value, parent_name=ws[f"D{row}"].value,
                              parent_number=ws[f"E{row}"].value, school=ws[f"F{row}"].value,
                              grade=ws[f"G{row}"].value, robotic_course=ws[f"H{row}"].value,
                              student_facilities=ws[f"J{row}"].value)
        students.append(new_student)
    return students


def create_grade_list(students):
    grade_list = dict()
    for student in students:
        try:
            grade_list[student.grade].append(student)
        except KeyError:
            grade_list[student.grade] = list()
            grade_list[student.grade].append(student)

    return grade_list


def basic_elimination(students):
    eliminated_classes = ["1", "2", "8"]
    grade_list = create_grade_list(students)

    for sinif in eliminated_classes:
        grade_list.pop(sinif, None)

    return grade_list


def draw_students_from_excel(dosya_adi):
    students = create_students_list(dosya_adi)
    sinif_kontenjanlari = dict()
    hak_kazananlar = list()

    grade_list = basic_elimination(students)

    for sinif in grade_list.keys():
        sinif_kontenjanlari[sinif] = 16

    for grade, students in grade_list.items():
        students = sorted(students, key=lambda i: i.facility_note(), reverse=True)
        grade_list[grade] = students

    for grade, students in grade_list.items():
        for student in students:
            if student.is_completed_course():
                students.remove(student)
                hak_kazananlar.append(student)
                sinif_kontenjanlari[student.grade] -= 1

        last_student = students[sinif_kontenjanlari[grade] - 1]
        after_last_student = students[sinif_kontenjanlari[grade]]
        if last_student.facility_note() == after_last_student.facility_note():
            i = sinif_kontenjanlari[grade] - 1
            while last_student.facility_note() == students[i].facility_note():
                i -= 1
            k = sinif_kontenjanlari[grade]
            while last_student.facility_note() == students[k].facility_note():
                k += 1

            sinif_kontenjanlari[grade] -= i + 1
            hak_kazananlar.extend(students[:i + 1])
            kura_listesi = list()
            equal_note = last_student.facility_note()
            for student in students[i + 1:k]:
                if student.facility_note():
                    kura_listesi.append(student)
            hak_kazananlar.extend(sample(kura_listesi, k=sinif_kontenjanlari[grade]))
        else:
            hak_kazananlar.extend(students[:sinif_kontenjanlari[grade]])
            sinif_kontenjanlari[grade] = 0

    hak_kazananlar = sorted(hak_kazananlar, key=lambda i: i.grade)
    return hak_kazananlar


def save_to_excel(grade_list: dict):
    wb = Workbook()

    for grade, students in grade_list.items():
        wb.create_sheet(f"{grade}. Sınıflar")
        sheet = wb.get_sheet_by_name(f"{grade}. Sınıflar")

        for num, student in enumerate(students, 1):
            sheet[f"A{num}"] = student.email
            sheet[f"B{num}"] = student.name
            sheet[f"C{num}"] = student.parent_name
            sheet[f"D{num}"] = student.parent_number
            sheet[f"F{num}"] = student.robotic_course
    timestr = time.strftime("%Y%m%d-%H%M%S")
    wb.save(f"sonuclar{timestr}.xlsx")
